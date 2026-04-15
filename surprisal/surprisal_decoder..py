from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from transformers import AutoTokenizer, AutoModelForCausalLM

import math
import pandas as pd
import torch
import torch.nn.functional as F

# Model Card

model_name = ""

# Load Model and Tokenizer

model = AutoModelForCausalLM.from_pretrained(model_name)
tokenizer = AutoTokenizer.from_pretrained(model_name)
model.eval()
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
model.to(device)

# Function to Compute Token-level Surprisal

def compute_token_surprisal(sentence):
    input_ids = tokenizer.encode(sentence, return_tensors = "pt").to(device)

    with torch.no_grad():
        outputs = model(input_ids)
        logits = outputs.logits[0]

    surprisals = []
    for i in range(1, len(input_ids[0])):
        probs = F.softmax(logits[i - 1], dim = -1)
        token_id = input_ids[0, i]
        prob = probs[token_id].item()
        surprisal = -torch.log2(torch.tensor(prob)).item()
        token_str = tokenizer.decode([token_id])
        surprisals.append((token_str, surprisal))

    return surprisals

# Test Sentences

sentences = [
    "Who do you wanna congratulate after the conference tomorrow?",
    "Who do you wanna apologize after the conference tomorrow?"
    ]

# Compute Surprisal

for sentence in sentences:
    print("\n" + "-" * 50)
    print(f"{'Sentence:':<10} {sentence}")
    print("-" * 50)

    surprisal_results = compute_token_surprisal(sentence)

    for token, surprisal in surprisal_results:
        print(f"{token:>10} : {surprisal:.5f}")

# Load Excel File

df = pd.read_excel(".xlsx")

def compute_last_token_surprisal(sentence):
    input_ids = tokenizer.encode(sentence, return_tensors="pt").to(device)

    with torch.no_grad():
        outputs = model(input_ids)
        logits = outputs.logits[0]  # shape: [seq_len, vocab_size]

    surprisals = []
    for i in range(1, input_ids.size(1)):
        probs = torch.nn.functional.softmax(logits[i - 1], dim=-1)
        token_id = input_ids[0, i]
        prob = probs[token_id].item()
        surprisal = -torch.log2(torch.tensor(prob)).item()
        surprisals.append(surprisal)

    # Return surprisal of the last token
    if surprisals:
        return round(surprisals[-2], 4)
    else:
        return None

  # Save Results

output_file = ".xlsx"

with pd.ExcelWriter(output_file, engine = "openpyxl") as writer:
    df.to_excel(writer, index = False, sheet_name = "Sheet1")
    ws = writer.book["Sheet1"]

    # Set Font
    font = Font(name = "Arial", size = 12)

    # Set Column Widths
    column_widths = {1: 20, 2: 70, 3: 15}
    for col_idx, width in column_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # Apply Font and Alignment
    for row_idx, row in enumerate(ws.iter_rows()):
        for col_idx, cell in enumerate(row):
            cell.font = font
            if row_idx == 0:
                # Header Row: Center Everything
                cell.alignment = Alignment(horizontal = "center", vertical = "center")
            else:
                if col_idx == 0:
                    cell.alignment = Alignment(horizontal = "center", vertical = "center")
                elif col_idx == 1:
                    cell.alignment = Alignment(horizontal = "left", vertical = "center")
                elif col_idx == 2:
                    cell.alignment = Alignment(horizontal = "right", vertical = "center")
